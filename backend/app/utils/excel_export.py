from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="1E40AF",
    end_color="1E40AF"
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True
)

THIN_BORDER = Border(

    left=Side(style="thin"),

    right=Side(style="thin"),

    top=Side(style="thin"),

    bottom=Side(style="thin")

)


def style_header(ws):

    for cell in ws[1]:

        cell.fill = HEADER_FILL

        cell.font = HEADER_FONT

        cell.border = THIN_BORDER

        cell.alignment = Alignment(
            horizontal="center"
        )

    ws.freeze_panes = "A2"


def autofit(ws):

    for column in ws.columns:

        length = 0

        letter = get_column_letter(
            column[0].column
        )

        for cell in column:

            try:

                length = max(
                    length,
                    len(str(cell.value))
                )

            except:
                pass

        ws.column_dimensions[
            letter
        ].width = min(length + 4, 60)


def export_ai_report(data):

    wb = Workbook()

    summary = wb.active

    summary.title = "Executive Summary"

    # ==========================
    # Executive Summary
    # ==========================

    summary.append([
        "Metric",
        "Value"
    ])

    style_header(summary)

    analytics = data.get("analytics", {})

    summary.append([
        "Total Requirements",
        len(data.get("requirements", []))
    ])

    summary.append([
        "Clarification Questions",
        len(data.get("questions", []))
    ])

    summary.append([
        "Generated Test Cases",
        len(data.get("testcases", []))
    ])

    summary.append([
        "Coverage",
        f"{analytics.get('coverage', 0)}%"
    ])

    summary.append([
        "Covered Requirements",
        analytics.get(
            "covered_requirements",
            0
        )
    ])

    summary.append([
        "Missing Requirements",
        analytics.get(
            "missing_requirements",
            0
        )
    ])

    summary.append([
        "High Priority",
        analytics.get(
            "priority_summary",
            {}
        ).get(
            "High",
            0
        )
    ])

    summary.append([
        "Medium Priority",
        analytics.get(
            "priority_summary",
            {}
        ).get(
            "Medium",
            0
        )
    ])

    summary.append([
        "Low Priority",
        analytics.get(
            "priority_summary",
            {}
        ).get(
            "Low",
            0
        )
    ])

    autofit(summary)

    # ==========================
    # Requirements Sheet
    # ==========================

    requirements_ws = wb.create_sheet(
        "Requirements"
    )

    requirements_ws.append([
        "Requirement ID",
        "Module",
        "Requirement"
    ])

    style_header(requirements_ws)

    for req in data.get("requirements", []):

        requirements_ws.append([

            req.get(
                "requirement_id",
                ""
            ),

            req.get(
                "module",
                ""
            ),

            req.get(
                "requirement",
                ""
            )

        ])

        row = requirements_ws.max_row

        for cell in requirements_ws[row]:

            cell.border = THIN_BORDER

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    autofit(requirements_ws)

    # ==========================
    # Clarifications Sheet
    # ==========================

    questions_ws = wb.create_sheet(
        "Clarifications"
    )

    questions_ws.append([
        "Requirement ID",
        "Clarification Question"
    ])

    style_header(questions_ws)

    for question in data.get("questions", []):

        questions_ws.append([

            question.get(
                "requirement_id",
                ""
            ),

            question.get(
                "question",
                ""
            )

        ])

        row = questions_ws.max_row

        for cell in questions_ws[row]:

            cell.border = THIN_BORDER

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    autofit(questions_ws)

    # ==========================
    # Test Cases Sheet
    # ==========================

    testcases_ws = wb.create_sheet(
        "Test Cases"
    )

    testcases_ws.append([
        "Test Case ID",
        "Requirement ID",
        "Module",
        "Scenario",
        "Priority",
        "Type",
        "Preconditions",
        "Overall Expected Result"
    ])

    style_header(testcases_ws)

    for tc in data.get("testcases", []):

        preconditions = "\n".join(
            tc.get("preconditions", [])
        )

        testcases_ws.append([

            tc.get(
                "testcase_id",
                ""
            ),

            tc.get(
                "requirement_id",
                ""
            ),

            tc.get(
                "module",
                ""
            ),

            tc.get(
                "scenario",
                ""
            ),

            tc.get(
                "priority",
                ""
            ),

            tc.get(
                "type",
                ""
            ),

            preconditions,

            tc.get(
                "overall_expected_result",
                ""
            )

        ])

        row = testcases_ws.max_row

        for cell in testcases_ws[row]:

            cell.border = THIN_BORDER

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    autofit(testcases_ws)

    # ==========================
    # Detailed Test Steps
    # ==========================

    steps_ws = wb.create_sheet(
        "Test Steps"
    )

    steps_ws.append([
        "Test Case ID",
        "Step No",
        "Actor",
        "Navigation",
        "Action",
        "Expected Result"
    ])

    style_header(steps_ws)

    for tc in data.get("testcases", []):

        for step in tc.get("steps", []):

            steps_ws.append([

                tc.get(
                    "testcase_id",
                    ""
                ),

                step.get(
                    "step_number",
                    ""
                ),

                step.get(
                    "actor",
                    ""
                ),

                step.get(
                    "navigation",
                    ""
                ),

                step.get(
                    "action",
                    ""
                ),

                step.get(
                    "expected_result",
                    ""
                )

            ])

            row = steps_ws.max_row

            for cell in steps_ws[row]:

                cell.border = THIN_BORDER

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

    autofit(steps_ws)

    # ==========================
    # Traceability Matrix
    # ==========================

    traceability_ws = wb.create_sheet(
        "Traceability Matrix"
    )

    traceability_ws.append([
        "Requirement ID",
        "Module",
        "Requirement",
        "Coverage",
        "Linked Test Cases",
        "Test Case Count"
    ])

    style_header(traceability_ws)

    for row in data.get("traceability", []):

        traceability_ws.append([

            row.get(
                "requirement_id",
                ""
            ),

            row.get(
                "module",
                ""
            ),

            row.get(
                "requirement",
                ""
            ),

            row.get(
                "coverage",
                ""
            ),

            ", ".join(
                row.get(
                    "linked_testcases",
                    []
                )
            ),

            row.get(
                "testcase_count",
                row.get("count", 0)
            )

        ])

        current = traceability_ws.max_row

        for cell in traceability_ws[current]:

            cell.border = THIN_BORDER

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    autofit(traceability_ws)

    # ==========================
    # Requirement Quality
    # ==========================

    quality_ws = wb.create_sheet(
        "Requirement Quality"
    )

    quality_ws.append([
        "Requirement ID",
        "Quality Score",
        "Issues"
    ])

    style_header(quality_ws)

    for item in data.get("quality", []):

        quality_ws.append([

            item.get(
                "requirement_id",
                ""
            ),

            item.get(
                "quality_score",
                ""
            ),

            "\n".join(
                item.get(
                    "issues",
                    []
                )
            )

        ])

        current = quality_ws.max_row

        for cell in quality_ws[current]:

            cell.border = THIN_BORDER

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    autofit(quality_ws)

    # ==========================
    # Save Workbook
    # ==========================

    wb.save(
        "generated_testcases.xlsx"
    )